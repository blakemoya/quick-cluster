import argparse
import os
from openpyxl import load_workbook
import numpy as np
import hypertools as hyp

parser = argparse.ArgumentParser(description='')
parser.add_argument('--filename', dest='filename', default='sample.xlsx', help='filename.xlsx')
parser.add_argument('--input_dir', dest='input_dir', default='input', help='name of input dir')
parser.add_argument('--output_dir', dest='output_dir', default='output', help='name of output dir')
parser.add_argument('--label_column', dest='label_column', type=int, default=1, help='column with datapoint names')
parser.add_argument('--columns', dest='columns', type=int, default=8, help='# of columns')
parser.add_argument('--start_col', dest='start_col', type=int, default=2, help='first column to read')
parser.add_argument('--rows', dest='rows', type=int, default=1000, help='# of rows')
parser.add_argument('--start_row', dest='start_row', type=int, default=2, help='first row to read')
parser.add_argument('--clusters', dest='clusters', type=int, default=10, help='number of groups agglomerator will find')
args = parser.parse_args()


def main(_, f=args.filename, id=args.input_dir, od=args.output_dir, lc=args.label_column, c=args.columns,
         sc=args.start_col, r=args.rows, sr=args.start_row, ec=args.clusters):

    # create directories
    if not os.path.exists(id):
        os.makedirs(id)
    if not os.path.exists(od):
        os.makedirs(od)

    # open spreadsheet
    wb = load_workbook(filename=id + '/' + f)
    ws = wb.active

    # copy specified rows and column range into a numpy array
    arr = []
    for i in range(sr, r + sr):
        arr.append([ws.cell(row=i, column=j).value for j in range(sc, c + sc)])
    arr = np.array(arr)

    # the actual clustering and TSNE operation
    arr_reduced = hyp.reduce(arr, reduce='TSNE', ndims=3)
    agg_labels = hyp.cluster(arr, n_clusters=ec, cluster='AgglomerativeClustering')

    # read agglomerated clusters into a numpy array
    labeled_arr = []
    for i in range(sr, r + sr):
        labeled_arr.append([agg_labels[i - sr], ws.cell(row=i, column=lc).value])
    labeled_arr = np.array(labeled_arr)
    sorted_labeled_arr = labeled_arr
    sorted_labeled_arr = sorted_labeled_arr[sorted_labeled_arr[:, 0].argsort()]

    # write real labels into agglomerated groups onto a new worksheet on a copy of the input workbook
    ws2 = wb.create_sheet('Agglomerative Labels')
    n = 0
    j = 1
    for row in sorted_labeled_arr:
        if row[0] == str(n):
            ws2.cell(row=j, column=(n + 1)).value = row[1]
            j = j + 1
        else:
            n = n + 1
            j = 1
    wb.save(od + '/output.xlsx')

    hyp.plot(arr_reduced, '.', normalize='within', hue=agg_labels, legend=True, labels=labeled_arr[:, 1],
             explore=True)


if __name__ == '__main__':
    main(args)
